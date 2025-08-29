import pandas as pd
from datetime import datetime
from typing import List
from .customer_agent import CustomerAgent
from .customer_model import CustomerModel
from config import CAMPAIGN_START, CAMPAIGN_END
import os
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

def generate_campaign_excel_report(model: CustomerModel, output_filename: str | None = None) -> str:
    """
    Generate an Excel report with campaign statistics and customer information.
    
    Args:
        model: The CustomerModel instance after simulation
        output_filename: Optional custom filename for the Excel file
    
    Returns:
        str: Path to the generated Excel file
    """
    
    # Create data directory if it doesn't exist
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    data_dir = os.path.join(project_root, "data")
    os.makedirs(data_dir, exist_ok=True)
    
    # Generate filename with timestamp if not provided
    if output_filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"campaign_report_{timestamp}.xlsx"
    
    output_path = os.path.join(data_dir, output_filename)
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # 1. Campaign Summary Sheet
        create_campaign_summary_sheet(model, writer)
        
        # 2. Prize Winners Sheet
        create_prize_winners_sheet(model, writer)
        
        # 3. Customer Performance Sheet
        create_customer_performance_sheet(model, writer)
        
        # 4. Daily Statistics Sheet
        create_daily_statistics_sheet(model, writer)
        
        # Apply modern styling to all sheets
        apply_modern_styling(writer)
    
    print(f"Excel report generated successfully: {output_path}")
    return output_path

def create_campaign_summary_sheet(model: CustomerModel, writer):
    """Create the campaign summary sheet with key metrics."""
    
    summary_data = {
        'Metric': [
            'Total Orders Received',
            'Total Revenue Generated',
            'New Customers Acquired',
            'Average Revenue per Order',
            'Campaign Start Date',
            'Campaign End Date',
            'Average Campaign Impact Factor',
            'Max Campaign Impact Factor',
            'Customers at Max Impact Factor'
        ],
        'Value': [
            model.received_orders_count,
            f"₾{model.generated_revenue:,.2f}",
            model.new_customers_count,
            f"₾{model.generated_revenue / model.received_orders_count:,.2f}" if model.received_orders_count > 0 else "₾0.00",
            CAMPAIGN_START.strftime("%Y-%m-%d"),
            CAMPAIGN_END.strftime("%Y-%m-%d"),
            f"{sum(agent.campaign_impact_factor for agent in model.agents if isinstance(agent, CustomerAgent)) / len([agent for agent in model.agents if isinstance(agent, CustomerAgent)]):.3f}",
            f"{max(agent.campaign_impact_factor for agent in model.agents if isinstance(agent, CustomerAgent)):.3f}",
            f"{len([agent for agent in model.agents if isinstance(agent, CustomerAgent) and agent.campaign_impact_factor >= 1.5])}"
        ]
    }
    
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name='Campaign Summary', index=False)
    
    # Apply modern styling to summary sheet
    worksheet = writer.sheets['Campaign Summary']
    apply_summary_sheet_styling(worksheet)

def create_prize_winners_sheet(model: CustomerModel, writer):
    """Create the prize winners sheet with customer prize information."""
    
    prize_winners_data = []
    
    for agent in model.agents:
        if isinstance(agent, CustomerAgent) and len(agent.prize_wins) > 0:
            prize_winners_data.append({
                'Customer ID': agent.customer_id,
                'Email': agent.email,
                'Prizes Won': ', '.join(agent.prize_wins) if agent.prize_wins else 'None',
                'Number of Prizes': len(agent.prize_wins),
                'Tickets Count': agent.tickets_count,
                'Campaign Impact Factor': agent.campaign_impact_factor,
                'Total Orders': agent.total_orders,
                'Historical Spending': f"₾{agent.historical_spending:,.2f}",
                'Average Order Value': f"₾{agent.avg_order_value:,.2f}",
                'Is New Customer': 'Yes' if agent.is_new_customer else 'No'
            })
    
    if prize_winners_data:
        df_prize_winners = pd.DataFrame(prize_winners_data)
        df_prize_winners.to_excel(writer, sheet_name='Prize Winners', index=False)
        
        # Apply modern styling to prize winners sheet
        worksheet = writer.sheets['Prize Winners']
        apply_data_sheet_styling(worksheet)
    else:
        # Create empty sheet with message
        empty_data = {'Message': ['No customers have won prizes in this campaign.']}
        df_empty = pd.DataFrame(empty_data)
        df_empty.to_excel(writer, sheet_name='Prize Winners', index=False)
        
        # Apply styling to empty sheet
        worksheet = writer.sheets['Prize Winners']
        apply_empty_sheet_styling(worksheet)

def create_customer_performance_sheet(model: CustomerModel, writer):
    """Create a comprehensive customer performance sheet."""
    
    customer_data = []
    
    for agent in model.agents:
        if isinstance(agent, CustomerAgent):
            customer_data.append({
                'Customer ID': agent.customer_id,
                'Email': agent.email,
                'Tickets Count': agent.tickets_count,
                'Campaign Impact Factor': agent.campaign_impact_factor,
                'Has Won Impact Factor': agent.hasWonImpactFactor,
                'Orders During Campaign': agent.new_order_count,
                'Prizes Won': ', '.join(agent.prize_wins) if agent.prize_wins else 'None',
                'Number of Prizes': len(agent.prize_wins),
                'Total Orders': agent.total_orders,
                'Historical Spending': f"₾{agent.historical_spending:,.2f}",
                'Average Order Value': f"₾{agent.avg_order_value:,.2f}",
                'Is New Customer': 'Yes' if agent.is_new_customer else 'No'
            })
    
    df_customers = pd.DataFrame(customer_data)
    df_customers.to_excel(writer, sheet_name='Customer Performance', index=False)
    
    # Apply modern styling to customer performance sheet
    worksheet = writer.sheets['Customer Performance']
    apply_data_sheet_styling(worksheet)

def create_daily_statistics_sheet(model: CustomerModel, writer):
    """Create a sheet with daily campaign statistics if available."""
    
    try:
        # Get the data collector dataframe
        data = model.datacollector.get_model_vars_dataframe()
        
        if not data.empty:
            # Reset index to get step numbers as a column
            data_reset = data.reset_index()
            data_reset.columns = ['Day', 'New Customers', 'Revenue', 'Orders']
            
            # Format revenue column
            data_reset['Revenue'] = data_reset['Revenue'].apply(lambda x: f"₾{x:,.2f}")
            
            data_reset.to_excel(writer, sheet_name='Daily Statistics', index=False)
            
            # Apply modern styling to daily statistics sheet
            worksheet = writer.sheets['Daily Statistics']
            apply_data_sheet_styling(worksheet)
        else:
            # Create empty sheet with message
            empty_data = {'Message': ['No daily statistics available for this campaign.']}
            df_empty = pd.DataFrame(empty_data)
            df_empty.to_excel(writer, sheet_name='Daily Statistics', index=False)
            
            # Apply styling to empty sheet
            worksheet = writer.sheets['Daily Statistics']
            apply_empty_sheet_styling(worksheet)
            
    except Exception as e:
        # Create empty sheet with error message
        error_data = {'Message': [f'Error retrieving daily statistics: {str(e)}']}
        df_error = pd.DataFrame(error_data)
        df_error.to_excel(writer, sheet_name='Daily Statistics', index=False)
        
        # Apply styling to error sheet
        worksheet = writer.sheets['Daily Statistics']
        apply_empty_sheet_styling(worksheet)

# =============================================================================
# STYLING FUNCTIONS
# =============================================================================

def apply_modern_styling(writer):
    """Apply modern styling to all sheets in the workbook."""
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        
        # Set default font for the entire sheet
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = Font(name='Segoe UI', size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center')

def apply_summary_sheet_styling(worksheet):
    """Apply modern styling to the campaign summary sheet."""
    
    # Define modern colors
    header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    header_font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
    metric_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    value_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # Style header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Style data rows
    for row in worksheet.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            cell.border = border
            if i == 0:  # Metric column
                cell.fill = metric_fill
                cell.font = Font(name='Segoe UI', size=10, bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # Value column
                cell.fill = value_fill
                cell.font = Font(name='Segoe UI', size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add some spacing
    worksheet.row_dimensions[1].height = 25

def apply_data_sheet_styling(worksheet):
    """Apply modern styling to data sheets (Prize Winners, Customer Performance, Daily Statistics)."""
    
    # Define modern colors
    header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    alt_row_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # Style header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Style data rows with alternating colors
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.border = border
            cell.font = Font(name='Segoe UI', size=9)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            else:
                cell.fill = alt_row_fill
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add some spacing
    worksheet.row_dimensions[1].height = 25

def apply_empty_sheet_styling(worksheet):
    """Apply modern styling to empty sheets with messages."""
    
    # Define modern colors
    header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    message_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    message_font = Font(name='Segoe UI', size=10, italic=True)
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # Style header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Style message row
    for cell in worksheet[2]:
        cell.fill = message_fill
        cell.font = message_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add some spacing
    worksheet.row_dimensions[1].height = 25
    worksheet.row_dimensions[2].height = 30
